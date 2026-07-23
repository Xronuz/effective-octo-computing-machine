"""
XAVFSIZ XONADON — Statistika xizmati.
Agregat so'rovlar, Excel/PDF eksport.
"""
import io
import logging
from datetime import date, datetime, timezone, timedelta
from typing import Optional, List

from sqlalchemy import select, func, case
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    user as user_model,
    hudud as hudud_model,
    muammo as muammo_model,
    audit as audit_model,
    lokatsiya as lokatsiya_model,
)
from app.schemas.statistika import (
    UmumiyStatistika,
    MuammoTuriStat,
    MuammoXavfStat,
    MuammoStatusStat,
    MFYStatistika,
    TopshiriqStat,
    IntizomStat,
    VaqtDavriStat,
    StatistikaResponse,
    XodimStatistika,
)

logger = logging.getLogger("xavfsiz_xonadon")


async def get_umumiy(db: AsyncSession) -> UmumiyStatistika:
    """Umumiy statistika — barcha asosiy ko'rsatkichlar."""
    # Xonadonlar soni
    xonadon_soni_res = await db.execute(select(func.count(hudud_model.Xonadon.id)))
    xonadon_soni = xonadon_soni_res.scalar() or 0

    # Barcha muammolar soni
    muammo_soni_res = await db.execute(select(func.count(muammo_model.Muammo.id)))
    muammo_soni = muammo_soni_res.scalar() or 0

    # Ochiq muammolar (ochiq + jarayonda)
    ochiq_res = await db.execute(
        select(func.count(muammo_model.Muammo.id))
        .where(muammo_model.Muammo.status.in_(["ochiq", "jarayonda"]))
    )
    ochiq_muammolar = ochiq_res.scalar() or 0

    # Yopilgan muammolar
    yopilgan_res = await db.execute(
        select(func.count(muammo_model.Muammo.id))
        .where(muammo_model.Muammo.status == "yopilgan")
    )
    yopilgan_muammolar = yopilgan_res.scalar() or 0

    # Faol xodimlar (holat == 'faol')
    xodim_res = await db.execute(
        select(func.count(user_model.User.id))
        .where(user_model.User.holat == user_model.UserStatus.faol)
    )
    xodim_soni = xodim_res.scalar() or 0

    # MFY lar soni
    mfy_res = await db.execute(select(func.count(hudud_model.Mfy.id)))
    mfy_soni = mfy_res.scalar() or 0

    # Kamida 1 marta tekshirilgan xonadonlar
    tekshirilgan_res = await db.execute(
        select(func.count(func.distinct(muammo_model.Muammo.xonadon_id)))
    )
    tekshirilgan = tekshirilgan_res.scalar() or 0

    foiz = round((tekshirilgan / xonadon_soni * 100) if xonadon_soni > 0 else 0, 1)

    return UmumiyStatistika(
        xonadon_soni=xonadon_soni,
        muammo_soni=muammo_soni,
        ochiq_muammolar=ochiq_muammolar,
        yopilgan_muammolar=yopilgan_muammolar,
        xodim_soni=xodim_soni,
        mfy_soni=mfy_soni,
        tekshirilgan_xonadon=tekshirilgan,
        foiz=foiz,
    )


async def get_muammo_turlari(db: AsyncSession) -> List[MuammoTuriStat]:
    """Muammo turlari bo'yicha taqsimot."""
    result = await db.execute(
        select(
            muammo_model.Muammo.turi,
            func.count(muammo_model.Muammo.id).label("soni"),
        )
        .group_by(muammo_model.Muammo.turi)
        .order_by(func.count(muammo_model.Muammo.id).desc())
    )
    return [MuammoTuriStat(turi=row.turi, soni=row.soni) for row in result]


async def get_muammo_xavf(db: AsyncSession) -> List[MuammoXavfStat]:
    """Xavf darajalari bo'yicha taqsimot."""
    result = await db.execute(
        select(
            muammo_model.Muammo.xavf,
            func.count(muammo_model.Muammo.id).label("soni"),
        )
        .group_by(muammo_model.Muammo.xavf)
        .order_by(
            case(
                (muammo_model.Muammo.xavf == "kritik", 1),
                (muammo_model.Muammo.xavf == "yuqori", 2),
                (muammo_model.Muammo.xavf == "orta", 3),
                else_=4,
            )
        )
    )
    return [MuammoXavfStat(xavf=row.xavf, soni=row.soni) for row in result]


async def get_muammo_status(db: AsyncSession) -> List[MuammoStatusStat]:
    """Status bo'yicha taqsimot."""
    result = await db.execute(
        select(
            muammo_model.Muammo.status,
            func.count(muammo_model.Muammo.id).label("soni"),
        )
        .group_by(muammo_model.Muammo.status)
    )
    return [MuammoStatusStat(status=row.status, soni=row.soni) for row in result]


async def get_mfy_statistika(db: AsyncSession) -> List[MFYStatistika]:
    """Har bir MFY bo'yicha statistikalar."""
    # Xonadonlar soni MFY bo'yicha
    xonadon_subq = (
        select(
            hudud_model.Kocha.mfy_id,
            func.count(hudud_model.Xonadon.id).label("cnt"),
        )
        .join(hudud_model.Xonadon, hudud_model.Xonadon.kocha_id == hudud_model.Kocha.id)
        .group_by(hudud_model.Kocha.mfy_id)
        .subquery()
    )

    # Ochiq muammolar MFY bo'yicha
    ochiq_subq = (
        select(
            muammo_model.Muammo.xonadon_id,
        )
        .where(muammo_model.Muammo.status.in_(["ochiq", "jarayonda"]))
        .subquery()
    )

    ochiq_mfy_subq = (
        select(
            hudud_model.Kocha.mfy_id,
            func.count(func.distinct(ochiq_subq.c.xonadon_id)).label("cnt"),
        )
        .join(hudud_model.Xonadon, hudud_model.Xonadon.kocha_id == hudud_model.Kocha.id)
        .join(ochiq_subq, ochiq_subq.c.xonadon_id == hudud_model.Xonadon.id)
        .group_by(hudud_model.Kocha.mfy_id)
        .subquery()
    )

    # Yopilgan muammolar MFY bo'yicha
    yopiq_subq = (
        select(
            muammo_model.Muammo.xonadon_id,
        )
        .where(muammo_model.Muammo.status == "yopilgan")
        .subquery()
    )

    yopiq_mfy_subq = (
        select(
            hudud_model.Kocha.mfy_id,
            func.count(func.distinct(yopiq_subq.c.xonadon_id)).label("cnt"),
        )
        .join(hudud_model.Xonadon, hudud_model.Xonadon.kocha_id == hudud_model.Kocha.id)
        .join(yopiq_subq, yopiq_subq.c.xonadon_id == hudud_model.Xonadon.id)
        .group_by(hudud_model.Kocha.mfy_id)
        .subquery()
    )

    # Tekshirilgan xonadonlar MFY bo'yicha (kamida 1 muammosi bor)
    tekshirilgan_subq = (
        select(
            hudud_model.Kocha.mfy_id,
            func.count(func.distinct(muammo_model.Muammo.xonadon_id)).label("cnt"),
        )
        .join(hudud_model.Xonadon, hudud_model.Xonadon.kocha_id == hudud_model.Kocha.id)
        .join(muammo_model.Muammo, muammo_model.Muammo.xonadon_id == hudud_model.Xonadon.id)
        .group_by(hudud_model.Kocha.mfy_id)
        .subquery()
    )

    # Asosiy so'rov
    result = await db.execute(
        select(
            hudud_model.Mfy.id,
            hudud_model.Mfy.nomi,
            hudud_model.Mfy.xonadon_soni,
            func.coalesce(xonadon_subq.c.cnt, 0).label("haqiqiy_xonadon"),
            func.coalesce(tekshirilgan_subq.c.cnt, 0).label("tekshirilgan"),
            func.coalesce(ochiq_mfy_subq.c.cnt, 0).label("ochiq"),
            func.coalesce(yopiq_mfy_subq.c.cnt, 0).label("yopilgan"),
        )
        .outerjoin(xonadon_subq, xonadon_subq.c.mfy_id == hudud_model.Mfy.id)
        .outerjoin(tekshirilgan_subq, tekshirilgan_subq.c.mfy_id == hudud_model.Mfy.id)
        .outerjoin(ochiq_mfy_subq, ochiq_mfy_subq.c.mfy_id == hudud_model.Mfy.id)
        .outerjoin(yopiq_mfy_subq, yopiq_mfy_subq.c.mfy_id == hudud_model.Mfy.id)
        .order_by(hudud_model.Mfy.raqami)
    )

    statlar = []
    for row in result:
        total = row.haqiqiy_xonadon or row.xonadon_soni
        foiz = round((row.tekshirilgan / total * 100) if total > 0 else 0, 1)
        statlar.append(
            MFYStatistika(
                mfy_id=row.id,
                mfy_nomi=row.nomi,
                xonadon_soni=row.haqiqiy_xonadon,
                tekshirilgan=row.tekshirilgan,
                ochiq_muammo=row.ochiq,
                yopilgan_muammo=row.yopilgan,
                foiz=foiz,
            )
        )
    return statlar


async def get_topshiriq_stat(db: AsyncSession) -> TopshiriqStat:
    """Topshiriqlar bo'yicha statistika."""
    jami_res = await db.execute(select(func.count(audit_model.Topshiriq.id)))
    jami = jami_res.scalar() or 0

    new_res = await db.execute(
        select(func.count(audit_model.Topshiriq.id))
        .where(audit_model.Topshiriq.status == "yangi")
    )
    yangi = new_res.scalar() or 0

    korildi_res = await db.execute(
        select(func.count(audit_model.Topshiriq.id))
        .where(audit_model.Topshiriq.status == "korildi")
    )
    korildi = korildi_res.scalar() or 0

    bajarildi_res = await db.execute(
        select(func.count(audit_model.Topshiriq.id))
        .where(audit_model.Topshiriq.status == "bajarildi")
    )
    bajarildi = bajarildi_res.scalar() or 0

    kechikkan_res = await db.execute(
        select(func.count(audit_model.Topshiriq.id))
        .where(audit_model.Topshiriq.status == "kechikkan")
    )
    kechikkan = kechikkan_res.scalar() or 0

    return TopshiriqStat(
        jami=jami,
        yangi=yangi,
        korildi=korildi,
        bajarildi=bajarildi,
        kechikkan=kechikkan,
    )


async def get_intizom_stat(db: AsyncSession) -> IntizomStat:
    """Intizom bo'yicha statistika."""
    jami_res = await db.execute(select(func.count(audit_model.Intizom.id)))
    jami = jami_res.scalar() or 0

    ogoh_res = await db.execute(
        select(func.count(audit_model.Intizom.id))
        .where(audit_model.Intizom.turi == "ogohlantirish")
    )
    ogohlantirish = ogoh_res.scalar() or 0

    hayfsan_res = await db.execute(
        select(func.count(audit_model.Intizom.id))
        .where(audit_model.Intizom.turi == "hayfsan")
    )
    hayfsan = hayfsan_res.scalar() or 0

    ragbat_res = await db.execute(
        select(func.count(audit_model.Intizom.id))
        .where(audit_model.Intizom.turi == "ragbat")
    )
    ragbat = ragbat_res.scalar() or 0

    return IntizomStat(
        jami=jami,
        ogohlantirish=ogohlantirish,
        hayfsan=hayfsan,
        ragbat=ragbat,
    )


async def get_vaqt_dinamika(
    db: AsyncSession,
    oylar: int = 12,
) -> List[VaqtDavriStat]:
    """Oxirgi N oy bo'yicha muammolar dinamikasi."""
    boshlanish = datetime.now(timezone.utc) - timedelta(days=30 * oylar)

    result = await db.execute(
        select(
            func.to_char(muammo_model.Muammo.sinxron_vaqti, "YYYY-MM").label("davr"),
            func.count(muammo_model.Muammo.id).label("ochilgan"),
            func.sum(
                case(
                    (muammo_model.Muammo.status == "yopilgan", 1),
                    else_=0,
                )
            ).label("yopilgan"),
        )
        .where(muammo_model.Muammo.sinxron_vaqti >= boshlanish)
        .group_by("davr")
        .order_by("davr")
    )
    return [
        VaqtDavriStat(
            davr=row.davr,
            ochilgan=row.ochilgan,
            yopilgan=row.yopilgan or 0,
        )
        for row in result
    ]


async def get_xodim_statistika(
    db: AsyncSession,
    page: int = 1,
    size: int = 20,
) -> tuple[List[XodimStatistika], int]:
    """Xodimlar bo'yicha statistika (sahifalangan)."""
    # Jami xodimlar soni
    total_res = await db.execute(
        select(func.count(user_model.User.id))
        .where(user_model.User.holat == user_model.UserStatus.faol)
    )
    total = total_res.scalar() or 0

    # Xodimlar ro'yxati + muammo sonlari
    result = await db.execute(
        select(
            user_model.User.id,
            user_model.User.familiya,
            user_model.User.ism,
            user_model.User.sharif,
            func.count(muammo_model.Muammo.id).label("jami_muammo"),
            func.sum(
                case(
                    (muammo_model.Muammo.status.in_(["ochiq", "jarayonda"]), 1),
                    else_=0,
                )
            ).label("ochiq_muammo"),
            func.sum(
                case(
                    (muammo_model.Muammo.status == "yopilgan", 1),
                    else_=0,
                )
            ).label("yopilgan_muammo"),
            func.max(muammo_model.Muammo.sinxron_vaqti).label("oxirgi_muammo"),
        )
        .outerjoin(muammo_model.Muammo, muammo_model.Muammo.xodim_id == user_model.User.id)
        .where(user_model.User.holat == user_model.UserStatus.faol)
        .group_by(user_model.User.id)
        .order_by(func.count(muammo_model.Muammo.id).desc())
        .offset((page - 1) * size)
        .limit(size)
    )

    xodimlar = []
    for row in result:
        fio = f"{row.familiya} {row.ism} {row.sharif or ''}".strip()
        oxirgi = row.oxirgi_muammo.isoformat() if row.oxirgi_muammo else None
        xodimlar.append(
            XodimStatistika(
                xodim_id=row.id,
                xodim_fio=fio,
                jami_muammo=row.jami_muammo or 0,
                ochiq_muammo=row.ochiq_muammo or 0,
                yopilgan_muammo=row.yopilgan_muammo or 0,
                jami_tekshirish=row.jami_muammo or 0,
                oxirgi_faollik=oxirgi,
            )
        )

    return xodimlar, total


async def get_full_statistika(db: AsyncSession) -> StatistikaResponse:
    """Barcha statistikani bir so'rovda yig'ish.

    So'rovlar ketma-ket bajariladi — bitta AsyncSession (asyncpg ulanishi)
    ustida parallel query ga ruxsat yo'q.
    """
    umumiy = await get_umumiy(db)
    turlar = await get_muammo_turlari(db)
    xavf = await get_muammo_xavf(db)
    status = await get_muammo_status(db)
    mfylar = await get_mfy_statistika(db)
    topshiriqlar = await get_topshiriq_stat(db)
    intizom = await get_intizom_stat(db)
    dinamika = await get_vaqt_dinamika(db)
    return StatistikaResponse(
        umumiy=umumiy,
        muammo_turlari=turlar,
        muammo_xavf=xavf,
        muammo_status=status,
        mfylar=mfylar,
        topshiriqlar=topshiriqlar,
        intizom=intizom,
        vaqt_dinamika=dinamika,
    )


# ============ Excel eksport ============

def generate_excel(statistika: StatistikaResponse) -> io.BytesIO:
    """Statistikani Excel faylga eksport qilish."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl o'rnatilmagan: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statistika"

    # Sarlavha uslubi
    header_font = Font(bold=True, size=12, color="1F2937")
    header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row = 1
    # Umumiy ko'rsatkichlar
    ws.cell(row, 1, "XAVFSIZ XONADON — STATISTIKA HISOBOTI").font = Font(bold=True, size=16)
    row += 2

    ws.cell(row, 1, "UMUMIY KO'RSATKICHLAR").font = header_font
    row += 1

    umumiy = statistika.umumiy
    fields = [
        ("Jami xonadonlar", umumiy.xonadon_soni),
        ("Jami muammolar", umumiy.muammo_soni),
        ("Ochiq muammolar", umumiy.ochiq_muammolar),
        ("Yopilgan muammolar", umumiy.yopilgan_muammolar),
        ("Faol xodimlar", umumiy.xodim_soni),
        ("MFY lar", umumiy.mfy_soni),
        ("Tekshirilgan xonadonlar", umumiy.tekshirilgan_xonadon),
        ("Tekshirilgan %", f"{umumiy.foiz}%"),
    ]
    for label, val in fields:
        ws.cell(row, 1, label).font = Font(bold=True)
        ws.cell(row, 2, val)
        row += 1

    row += 1

    # MFY bo'yicha
    ws.cell(row, 1, "MFY BO'YICHA").font = header_font
    row += 1
    mfy_headers = ["MFY", "Xonadon", "Tekshirilgan", "Ochiq muammo", "Yopilgan", "Foiz"]
    for col_idx, h in enumerate(mfy_headers, 1):
        cell = ws.cell(row, col_idx, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    for m in statistika.mfylar:
        ws.cell(row, 1, m.mfy_nomi).border = thin_border
        ws.cell(row, 2, m.xonadon_soni).border = thin_border
        ws.cell(row, 3, m.tekshirilgan).border = thin_border
        ws.cell(row, 4, m.ochiq_muammo).border = thin_border
        ws.cell(row, 5, m.yopilgan_muammo).border = thin_border
        ws.cell(row, 6, f"{m.foiz}%").border = thin_border
        row += 1

    row += 1

    # Topshiriqlar
    ws.cell(row, 1, "TOPSHIRIQLAR").font = header_font
    row += 1
    t = statistika.topshiriqlar
    for label, val in [
        ("Jami", t.jami), ("Yangi", t.yangi), ("Ko'rilgan", t.korildi),
        ("Bajarilgan", t.bajarildi), ("Kechikkan", t.kechikkan),
    ]:
        ws.cell(row, 1, label).font = Font(bold=True)
        ws.cell(row, 2, val)
        row += 1

    row += 1

    # Intizom
    ws.cell(row, 1, "INTIZOM").font = header_font
    row += 1
    i = statistika.intizom
    for label, val in [
        ("Jami", i.jami), ("Ogohlantirish", i.ogohlantirish),
        ("Hayfsan", i.hayfsan), ("Rag'bat", i.ragbat),
    ]:
        ws.cell(row, 1, label).font = Font(bold=True)
        ws.cell(row, 2, val)
        row += 1

    # Ustun kengligi
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============ PDF eksport ============

def generate_pdf(statistika: StatistikaResponse) -> io.BytesIO:
    """Statistikani PDF faylga eksport qilish."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        )
        from reportlab.lib.units import mm
    except ImportError:
        raise ImportError("reportlab o'rnatilmagan: pip install reportlab")

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = []

    # Sarlavha
    title_style = ParagraphStyle(
        "Title2", parent=styles["Title"], fontSize=16, spaceAfter=10, textColor=HexColor("#1F2937")
    )
    elements.append(Paragraph("XAVFSIZ XONADON — Statistik hisobot", title_style))
    elements.append(Spacer(1, 5 * mm))

    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"], fontSize=13, spaceBefore=8, spaceAfter=4,
        textColor=HexColor("#374151"),
    )

    # Umumiy
    elements.append(Paragraph("UMUMIY KO'RSATKICHLAR", section_style))
    u = statistika.umumiy
    data = [
        ["Jami xonadonlar", str(u.xonadon_soni)],
        ["Jami muammolar", str(u.muammo_soni)],
        ["Ochiq muammolar", str(u.ochiq_muammolar)],
        ["Yopilgan muammolar", str(u.yopilgan_muammolar)],
        ["Faol xodimlar", str(u.xodim_soni)],
        ["MFY lar soni", str(u.mfy_soni)],
        ["Tekshirilgan", str(u.tekshirilgan_xonadon)],
        ["Tekshirilgan %", f"{u.foiz}%"],
    ]
    t = Table(data, colWidths=[140, 60])
    t.setStyle(
        TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, HexColor("#D1D5DB")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    elements.append(t)
    elements.append(Spacer(1, 5 * mm))

    # Topshiriqlar
    elements.append(Paragraph("TOPSHIRIQLAR", section_style))
    tp = statistika.topshiriqlar
    data2 = [
        ["Jami", str(tp.jami)], ["Yangi", str(tp.yangi)],
        ["Ko'rilgan", str(tp.korildi)], ["Bajarilgan", str(tp.bajarildi)],
        ["Kechikkan", str(tp.kechikkan)],
    ]
    t2 = Table(data2, colWidths=[140, 60])
    t2.setStyle(
        TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, HexColor("#D1D5DB")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    elements.append(t2)
    elements.append(Spacer(1, 5 * mm))

    # Intizom
    elements.append(Paragraph("INTIZOM", section_style))
    iz = statistika.intizom
    data3 = [
        ["Jami", str(iz.jami)], ["Ogohlantirish", str(iz.ogohlantirish)],
        ["Hayfsan", str(iz.hayfsan)], ["Rag'bat", str(iz.ragbat)],
    ]
    t3 = Table(data3, colWidths=[140, 60])
    t3.setStyle(
        TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, HexColor("#D1D5DB")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    elements.append(t3)

    doc.build(elements)
    output.seek(0)
    return output
